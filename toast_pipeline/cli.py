"""CLI entry point.

  python -m toast_pipeline.cli init-db
  python -m toast_pipeline.cli run --start 2026-06-01 --end 2026-06-07 [--locations BALLPARK,MVT]
  python -m toast_pipeline.cli run            # self-healing window: yesterday back-padded 2 days
  python -m toast_pipeline.cli validate
  python -m toast_pipeline.cli bikky-instore  # load all P*IS.csv from Data/Bikkydata/InStore/

Stage order per run (mirrors the reference architecture):
  (0) truncate staging -> (1) config fetch (HALT on missing dining options)
  -> (2) order pull, raw landing -> (3) parse + clean -> staging
  -> (4) merge to public -> (5) validate counts
"""
from __future__ import annotations

import argparse
import csv
import logging
import re
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from . import config, db
from .fetch import config_api, orders as orders_fetch
from .parse.orders import parse_order

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(name)s %(levelname)s %(message)s")
log = logging.getLogger("cli")

# Locations permanently closed / no longer reachable via Toast's live API.
# Excluded from the default `run` (scheduled pull) only — `reparse` and direct
# DB work still resolve these via config.load_locations() since they don't
# call the live API, so historical data stays fully manipulable.
CLOSED_LOCATIONS = {"BALLPARK"}


def _pull_location(loc: config.Location, run_id: int, start: date, end: date) -> dict:
    """One worker per location: own connection, own lookups, own batch."""
    conn = db.connect()
    counts = {"fetched": 0, "landed": 0, "lines": 0}
    try:
        cfg = config_api.fetch_all_config(loc)
        db.land_raw_config(conn, run_id, loc.code, cfg)
        lookups = config_api.build_lookups(cfg)
        # Sale-time-aware menu resolution (MENU_SALETIME_RESOLUTION_SPEC.md) —
        # snapshots now includes today's, just landed above.
        snapshots = db.fetch_menu_snapshots(conn, loc.code)
        lookups["menu_resolver"] = config_api.build_time_lookups(snapshots, cfg)

        batch = {k: [] for k in ("order_lines", "modifiers", "checks", "payments", "adjustments", "order_refunds")}
        page: list[dict] = []

        def _flush_page() -> None:
            if not page:
                return
            counts["landed"] += db.land_raw_orders_batch(conn, run_id, loc.code, page)
            for o in page:
                parsed = parse_order(o, loc.code, lookups)
                batch["order_lines"].extend(parsed.lines)
                batch["modifiers"].extend(parsed.modifiers)
                batch["checks"].extend(parsed.checks)
                batch["payments"].extend(parsed.payments)
                batch["adjustments"].extend(parsed.adjustments)
                batch["order_refunds"].extend(parsed.order_refunds)
            page.clear()

        for order in orders_fetch.fetch_orders(loc, start, end):
            counts["fetched"] += 1
            page.append(order)
            if len(page) >= 100:
                _flush_page()
        _flush_page()
        conn.commit()

        for kind, rows in batch.items():
            n = db.bulk_stage(conn, kind, rows)
            log.info("%s: staged %d %s", loc.code, n, kind)
        counts["lines"] = len(batch["order_lines"])
        return counts
    finally:
        conn.close()


def cmd_run(args: argparse.Namespace) -> None:
    locs = config.load_locations()
    if args.locations:
        wanted = {c.strip().upper() for c in args.locations.split(",")}
        locs = [l for l in locs if l.code.upper() in wanted]
    else:
        locs = [l for l in locs if l.code.upper() not in CLOSED_LOCATIONS]

    end = date.fromisoformat(args.end) if args.end else date.today() - timedelta(days=1)
    start = (date.fromisoformat(args.start) if args.start
             else end - timedelta(days=config.DEFAULT_BACKPAD_DAYS))
    if (end - start).days > config.MAX_WINDOW_DAYS:
        raise SystemExit(f"window exceeds {config.MAX_WINDOW_DAYS} days — split the backfill")

    conn = db.connect()
    db.truncate_staging(conn)
    run_id = db.open_pull_run(conn, start, end, [l.code for l in locs])
    log.info("pull_run %d: %s -> %s for %s", run_id, start, end, [l.code for l in locs])

    fetched = landed = 0
    try:
        with ThreadPoolExecutor(max_workers=len(locs)) as ex:
            for counts in ex.map(lambda l: _pull_location(l, run_id, start, end), locs):
                fetched += counts["fetched"]
                landed += counts["landed"]

        db.merge_to_public(conn)
        db.close_pull_run(conn, run_id, "success", fetched, landed)
        log.info("run %d complete: fetched=%d landed=%d", run_id, fetched, landed)
        cmd_validate(args)
        cmd_consolidate_names(args)
    except Exception as e:
        db.close_pull_run(conn, run_id, "failed", fetched, landed, error=str(e))
        raise
    finally:
        conn.close()


def cmd_init_db(args: argparse.Namespace) -> None:
    conn = db.connect()
    db.init_schema(conn)
    conn.close()
    log.info("schema initialized")


def cmd_reparse(args: argparse.Namespace) -> None:
    """Rebuild staging from the raw payloads already in the database — no
    Toast API calls. Use after changing mappings or cleaning rules, then
    follow with `merge` (or let this command do both with --merge)."""
    locs = config.load_locations()
    if args.locations:
        wanted = {c.strip().upper() for c in args.locations.split(",")}
        locs = [l for l in locs if l.code.upper() in wanted]
    start = date.fromisoformat(args.start)
    end = date.fromisoformat(args.end)

    conn = db.connect()
    db.truncate_staging(conn)
    total = 0
    for loc in locs:
        cfg = db.fetch_latest_config(conn, loc.code)
        if not cfg.get("dining_options"):
            log.warning("%s: no stored config found — skipping (run a pull first)", loc.code)
            continue
        lookups = config_api.build_lookups(cfg)
        # Sale-time-aware menu resolution — reparse output must not depend on
        # *when* the reparse is run (MENU_SALETIME_RESOLUTION_SPEC.md).
        snapshots = db.fetch_menu_snapshots(conn, loc.code)
        lookups["menu_resolver"] = config_api.build_time_lookups(snapshots, cfg)
        batch = {k: [] for k in ("order_lines", "modifiers", "checks", "payments", "adjustments", "order_refunds")}
        n = 0
        for payload in db.fetch_raw_orders(conn, loc.code, start, end):
            parsed = parse_order(payload, loc.code, lookups)
            batch["order_lines"].extend(parsed.lines)
            batch["modifiers"].extend(parsed.modifiers)
            batch["checks"].extend(parsed.checks)
            batch["payments"].extend(parsed.payments)
            batch["adjustments"].extend(parsed.adjustments)
            batch["order_refunds"].extend(parsed.order_refunds)
            n += 1
        for kind, rows in batch.items():
            db.bulk_stage(conn, kind, rows)
        log.info("%s: reparsed %d raw orders -> %d lines", loc.code, n, len(batch["order_lines"]))
        total += n
    if total and getattr(args, "merge", False):
        db.merge_to_public(conn)
        log.info("merge complete")
    conn.close()
    if total and getattr(args, "merge", False):
        cmd_validate(args)
        cmd_consolidate_names(args)


def cmd_precompute(args: argparse.Namespace) -> None:
    """Rebuild the dashboard's precomputed modifier-cost layer only (analytics.pc_*).
    Normally unnecessary — every merge refreshes it — but useful after editing
    sql/pc_refresh.sql or analytics source tables (r365 loads, overrides) directly."""
    conn = db.connect()
    db.refresh_precomputed(conn)
    conn.close()
    log.info("precomputed layer refreshed")


def cmd_merge(args: argparse.Namespace) -> None:
    """Merge whatever is currently in staging into public, then validate.
    Useful to finish a run whose pull succeeded but whose merge failed."""
    conn = db.connect()
    db.merge_to_public(conn)
    conn.close()
    log.info("merge complete")
    cmd_validate(args)
    cmd_consolidate_names(args)


def cmd_validate(args: argparse.Namespace) -> None:
    conn = db.connect()
    checks = {
        "staging lines": "SELECT count(*) FROM staging.order_lines",
        "public lines": "SELECT count(*) FROM public.fact_order_lines",
        "staging revenue (non-void)":
            "SELECT round(coalesce(sum(line_total),0),2) FROM staging.order_lines WHERE NOT is_voided",
        "orphan modifiers":
            "SELECT count(*) FROM public.fact_modifiers fm "
            "LEFT JOIN public.fact_order_lines fol ON fol.selection_guid = fm.parent_selection "
            "WHERE fol.selection_guid IS NULL",
        "lines missing channel":
            "SELECT count(*) FROM public.fact_order_lines WHERE channel_code IS NULL",
    }
    for label, q in checks.items():
        val = conn.execute(q).fetchone()[0]
        log.info("validate | %-28s %s", label, val)
    conn.close()


_BIKKY_DATA_ROOT = Path(__file__).resolve().parents[1] / "Data" / "Bikkydata"

_BIKKY_COL_MAP = {
    "Item":                               "item_name",
    "Item id":                            "item_id",
    "Item revenue":                       "revenue",
    "Item revenue per location":          "revenue_per_loc",
    "Item revenue percentage":            "revenue_pct",
    "Item volume":                        "volume",
    "Item volume per location":           "volume_per_loc",
    "Item volume percentage":             "volume_pct",
    "Item aov":                           "aov",
    "Item guests":                        "guests",
    "N day item return rate":             "return_rate",
    "N day item reorder rate":            "reorder_rate",
    "Business date previous start":       "prev_period_start",
    "Business date previous end":         "prev_period_end",
    "Item revenue previous":              "revenue_prev",
    "Item revenue per location previous": "revenue_per_loc_prev",
    "Item revenue percentage previous":   "revenue_pct_prev",
    "Item volume previous":               "volume_prev",
    "Item volume per location previous":  "volume_per_loc_prev",
    "Item volume percentage previous":    "volume_pct_prev",
    "Item aov previous":                  "aov_prev",
    "Item guests previous":               "guests_prev",
    "N day item return rate previous":    "return_rate_prev",
    "N day item reorder rate previous":   "reorder_rate_prev",
}

_BIKKY_DATE_COLS    = {"prev_period_start", "prev_period_end"}
_BIKKY_NUMERIC_COLS = {
    "revenue", "revenue_per_loc", "revenue_pct",
    "volume", "volume_per_loc", "volume_pct",
    "aov", "guests", "return_rate", "reorder_rate",
    "revenue_prev", "revenue_per_loc_prev", "revenue_pct_prev",
    "volume_prev", "volume_per_loc_prev", "volume_pct_prev",
    "aov_prev", "guests_prev", "return_rate_prev", "reorder_rate_prev",
}

_BIKKY_UPSERT_TMPL = """
    INSERT INTO {table} (
        fiscal_year, period, item_name, item_id,
        revenue, revenue_per_loc, revenue_pct,
        volume, volume_per_loc, volume_pct,
        aov, guests, return_rate, reorder_rate,
        prev_period_start, prev_period_end,
        revenue_prev, revenue_per_loc_prev, revenue_pct_prev,
        volume_prev, volume_per_loc_prev, volume_pct_prev,
        aov_prev, guests_prev, return_rate_prev, reorder_rate_prev
    ) VALUES (
        %(fiscal_year)s, %(period)s, %(item_name)s, %(item_id)s,
        %(revenue)s, %(revenue_per_loc)s, %(revenue_pct)s,
        %(volume)s, %(volume_per_loc)s, %(volume_pct)s,
        %(aov)s, %(guests)s, %(return_rate)s, %(reorder_rate)s,
        %(prev_period_start)s, %(prev_period_end)s,
        %(revenue_prev)s, %(revenue_per_loc_prev)s, %(revenue_pct_prev)s,
        %(volume_prev)s, %(volume_per_loc_prev)s, %(volume_pct_prev)s,
        %(aov_prev)s, %(guests_prev)s, %(return_rate_prev)s, %(reorder_rate_prev)s
    )
    ON CONFLICT (fiscal_year, period, item_name) DO UPDATE SET
        item_id              = EXCLUDED.item_id,
        revenue              = EXCLUDED.revenue,
        revenue_per_loc      = EXCLUDED.revenue_per_loc,
        revenue_pct          = EXCLUDED.revenue_pct,
        volume               = EXCLUDED.volume,
        volume_per_loc       = EXCLUDED.volume_per_loc,
        volume_pct           = EXCLUDED.volume_pct,
        aov                  = EXCLUDED.aov,
        guests               = EXCLUDED.guests,
        return_rate          = EXCLUDED.return_rate,
        reorder_rate         = EXCLUDED.reorder_rate,
        prev_period_start    = EXCLUDED.prev_period_start,
        prev_period_end      = EXCLUDED.prev_period_end,
        revenue_prev         = EXCLUDED.revenue_prev,
        revenue_per_loc_prev = EXCLUDED.revenue_per_loc_prev,
        revenue_pct_prev     = EXCLUDED.revenue_pct_prev,
        volume_prev          = EXCLUDED.volume_prev,
        volume_per_loc_prev  = EXCLUDED.volume_per_loc_prev,
        volume_pct_prev      = EXCLUDED.volume_pct_prev,
        aov_prev             = EXCLUDED.aov_prev,
        guests_prev          = EXCLUDED.guests_prev,
        return_rate_prev     = EXCLUDED.return_rate_prev,
        reorder_rate_prev    = EXCLUDED.reorder_rate_prev,
        loaded_at            = now()
"""


def _bikky_coerce(val: str, col: str):
    v = val.strip()
    if not v:
        return None
    if col in _BIKKY_DATE_COLS:
        return date.fromisoformat(v)
    if col in _BIKKY_NUMERIC_COLS:
        try:
            return Decimal(v)
        except InvalidOperation:
            return None
    return v


def _load_bikky_dir(data_dir: Path, glob: str, period_pattern: str,
                    sql_file: str, table: str, label: str) -> None:
    files = sorted(data_dir.glob(glob))
    if not files:
        raise SystemExit(f"No {glob} files found in {data_dir}")

    conn = db.connect()
    sql_path = Path(__file__).resolve().parents[1] / "sql" / sql_file
    conn.execute(sql_path.read_text())
    conn.commit()

    upsert = _BIKKY_UPSERT_TMPL.format(table=table)
    for path in files:
        m = re.match(period_pattern, path.stem, re.IGNORECASE)
        if not m:
            log.warning("skipping %s — can't parse period/year from filename", path.name)
            continue
        period = int(m.group(1))
        fiscal_year = int(m.group(2))

        rows = []
        with path.open(newline="", encoding="utf-8-sig") as f:
            for raw in csv.DictReader(f):
                row: dict = {"fiscal_year": fiscal_year, "period": period}
                for csv_col, db_col in _BIKKY_COL_MAP.items():
                    row[db_col] = _bikky_coerce(raw.get(csv_col, ""), db_col)
                if row.get("item_name"):
                    rows.append(row)

        if rows:
            with conn.cursor() as cur:
                cur.executemany(upsert, rows)
            conn.commit()
        log.info("%s: %s period=%d year=%d → %d rows upserted",
                 label, path.name, period, fiscal_year, len(rows))

    conn.close()


def cmd_bikky_instore(args: argparse.Namespace) -> None:
    _load_bikky_dir(
        data_dir=_BIKKY_DATA_ROOT / "InStore",
        glob="P*IS.csv",
        period_pattern=r"P(\d{2})(\d{4})IS",
        sql_file="006_bikky_instore.sql",
        table="public.fact_bikky_instore",
        label="bikky-instore",
    )


def cmd_bikky_3pd(args: argparse.Namespace) -> None:
    _load_bikky_dir(
        data_dir=_BIKKY_DATA_ROOT / "3PD+Loyalty",
        glob="P*Del.csv",
        period_pattern=r"P(\d{2})(\d{4})Del",
        sql_file="007_bikky_3pd_loyalty.sql",
        table="public.fact_bikky_3pd_loyalty",
        label="bikky-3pd",
    )


_LOOKUP_DIR = Path(__file__).resolve().parents[1] / "Data" / "LookupData"
_R365_DATA_ROOT = Path(__file__).resolve().parents[1] / "Data" / "R365Data"


def cmd_load_lookups(args: argparse.Namespace) -> None:
    """Load all lookup Excel files into the analytics schema — no data dropped."""
    import openpyxl

    conn = db.connect()
    sql_root = Path(__file__).resolve().parents[1] / "sql"

    for sql_file in ["008_analytics_modifier_type.sql", "009_analytics_parent_item_type.sql",
                     "010_analytics_item_lookup.sql"]:
        conn.execute((sql_root / sql_file).read_text())
    conn.commit()

    # -------------------------------------------------------------------------
    # LookupItemAndModifierType.xlsx
    # Section F-H (cols 5-7): modifier_name + item_type → modifier_type
    # Section J-K (cols 9-10): parent_item + item_type
    # -------------------------------------------------------------------------
    wb = openpyxl.load_workbook(_LOOKUP_DIR / "LookupItemAndModifierType.xlsx")
    ws = wb["Sheet1"]

    modifier_rows, parent_rows = [], []
    for row in ws.iter_rows(min_row=3, values_only=True):
        mod_name  = str(row[5]).strip() if row[5] else None
        item_type = str(row[6]).strip() if row[6] else None
        mod_type  = str(row[7]).strip() if row[7] else None
        if mod_name and item_type:
            modifier_rows.append((mod_name, item_type, mod_type))

        parent_item      = str(row[9]).strip()  if row[9]  else None
        parent_item_type = str(row[10]).strip() if row[10] else None
        if parent_item and parent_item_type:
            parent_rows.append((parent_item, parent_item_type))

    with conn.cursor() as cur:
        cur.executemany(
            """
            INSERT INTO analytics.modifier_type (modifier_name, item_type, modifier_type)
            VALUES (%s, %s, %s)
            ON CONFLICT (modifier_name, item_type) DO UPDATE SET
                modifier_type = EXCLUDED.modifier_type,
                loaded_at     = now()
            """,
            modifier_rows,
        )
        cur.executemany(
            """
            INSERT INTO analytics.parent_item_type (parent_item, item_type)
            VALUES (%s, %s)
            ON CONFLICT (parent_item, item_type) DO NOTHING
            """,
            parent_rows,
        )
    conn.commit()
    log.info("load-lookups: analytics.modifier_type → %d rows upserted", len(modifier_rows))
    log.info("load-lookups: analytics.parent_item_type → %d rows upserted", len(parent_rows))

    # -------------------------------------------------------------------------
    # LookupMenuBreakdown.xlsx → single analytics.item_lookup table
    # Section A (cols 0-1): raw_item_name → cleaned_item_name
    # Section B (cols 3-4): cleaned_item_name → category_1
    # Section C (cols 6-7): category_1 → category_2
    # Items in section B with no raw→cleaned entry are included as raw=cleaned.
    # -------------------------------------------------------------------------
    wb2 = openpyxl.load_workbook(_LOOKUP_DIR / "LookupMenuBreakdown.xlsx")
    ws2 = wb2["Sheet1"]

    raw_to_cleaned: dict[str, str] = {}
    cat1_map: dict[str, str] = {}
    cat2_map: dict[str, str] = {}

    for row in ws2.iter_rows(min_row=3, values_only=True):
        if row[0] and row[1]:
            raw_to_cleaned[str(row[0]).strip()] = str(row[1]).strip()
        if row[3] and row[4]:
            cat1_map[str(row[3]).strip()] = str(row[4]).strip()
        if row[6] and row[7]:
            cat2_map[str(row[6]).strip()] = str(row[7]).strip()

    item_lookup: dict[str, tuple] = {}
    for raw, cleaned in raw_to_cleaned.items():
        cat1 = cat1_map.get(cleaned)
        item_lookup[raw] = (raw, cleaned, cat1, cat2_map.get(cat1) if cat1 else None)

    covered_cleaned = set(raw_to_cleaned.values())
    for cleaned, cat1 in cat1_map.items():
        if cleaned not in covered_cleaned:
            item_lookup[cleaned] = (cleaned, cleaned, cat1, cat2_map.get(cat1))

    item_lookup_rows = list(item_lookup.values())

    with conn.cursor() as cur:
        cur.executemany(
            """
            INSERT INTO analytics.item_lookup
                (raw_item_name, cleaned_item_name, category_1, category_2)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (raw_item_name) DO UPDATE SET
                cleaned_item_name = EXCLUDED.cleaned_item_name,
                category_1        = EXCLUDED.category_1,
                category_2        = EXCLUDED.category_2,
                loaded_at         = now()
            """,
            item_lookup_rows,
        )
    conn.commit()
    log.info("load-lookups: analytics.item_lookup → %d rows upserted", len(item_lookup_rows))
    conn.close()


def _derive_modifier_clean_name(recipe_name: str) -> str:
    """Column B fallback when the sheet omits it — mirrors the spreadsheet
    formula =TRIM(SUBSTITUTE(A2,"MI ",""))). Excel's TRIM also collapses
    internal whitespace runs, not just leading/trailing, hence the regex."""
    return re.sub(r"\s+", " ", recipe_name.replace("MI ", "")).strip()


# recipe_name -> the clean_name it should actually have, overriding whatever the
# sheet says. R365 has repeatedly mis-exported column B for these SAME recipes
# across periods (wrong word order, missing/extra words, an accented character
# Toast's own modifier names don't use, etc.) — each was verified against real
# order data or fact_modifiers.canonical_name before being added here, not just
# guessed from naming similarity. See MODIFIER_COST_KNOWN_FIXES.md for the full
# rationale per entry and how to verify a new one before adding it.
_KNOWN_CLEAN_NAME_FIXES: dict[str, str] = {
    "MI Romaine": "Romaine Lettuce",
    "MI 1/2 Romaine": "1/2 Romaine Lettuce",
    "MI Romaine - Classic": "Romaine Lettuce - Classic",
    "MI Romaine - Desi Deluxe": "Romaine Lettuce - Desi Deluxe",
    "MI Romaine - Party Pack": "Romaine Lettuce - Party Pack",
    "MI Shredded Romaine": "Romaine",
    "MI 1/2 Harvest Veggies": "1/2 Roasted Vegetables",
    "MI Kokum Vinaigrette Dressing": "Kokum Vinaigrette",
    "MI Tikka Masala Sauce": "Tikka Masala",
    "MI 1/2 Chicken Tikka": "1/2 Chicken",
    "MI Sautéed Spinach": "Sauteed Spinach",
    "MI Sautéed Spinach - Classic": "Sauteed Spinach - Classic",
    "MI Sautéed Spinach - Desi Deluxe": "Sauteed Spinach - Desi Deluxe",
    "MI Sautéed Spinach - Party Pack": "Sauteed Spinach - Party Pack",
    "MI Sautéed Spinach - Side": "Sauteed Spinach - Side",
    "MI HUNGRY Sautéed Spinach": "HUNGRY Sauteed Spinach",
    "MI That Fire Hot Sauce - Bottle": "That Fire Hot Sauce (Bottle)",
    # Below: bulk-added from Rahul's modifier/catering reconciliation lists
    # (MODIFIER_COST_KNOWN_FIXES.md). Two categories were deliberately held back:
    #   1. 13 recipe_names that would need two different clean_names depending on
    #      context (e.g. "MI Lamb Kebab Meatballs" -> both "Lamb Kebab" and "Extra
    #      Lamb Kebab") — a single recipe row can only hold one clean_name.
    #   2. 23 recipe_names whose clean_name is ALSO a real modifier in the current
    #      IH/Online dashboard (e.g. "MI Basmati Rice" -> "Basmati Rice", used in
    #      BYO Grain Bowl today) — scope for now is catering/catering-3pd/offsite
    #      only, per explicit instruction; don't auto-touch anything the live
    #      non-catering dashboard already reads. From P7 onward, ask before adding
    #      any new fix outside that catering scope.
    # Both categories need a human call and are surfaced instead via the
    # r365-modifier-cost review-flag mechanism below.
    "MI Arugula - Classic": "Arugula - Classic",
    "MI Arugula - Party Pack": "Arugula - Party Pack",
    "MI Baby Spinach - Classic": "Baby Spinach - Classic",
    "MI Baby Spinach - Party Pack": "Baby Spinach - Party Pack",
    "MI Basmati Rice - Catering - Additional Item": "Catering - Additional Item - Basmati Rice",
    "MI Basmati Rice - Classic": "Basmati Rice - Classic",
    "MI Basmati Rice - Party Pack": "Basmati Rice - Party Pack",
    "MI Butter Chicken Burrito - In House": "Butter Chicken Burrito",
    "MI Carrot Slaw - Classic": "Carrot Slaw - Classic",
    "MI Carrot Slaw - Party Pack": "Carrot Slaw - Party Pack",
    "MI Cauliflower + Potato - Catering - Additional Item": "Catering - Additional Item - Cauliflower + Potato",
    "MI Cauliflower + Potato - Classic": "Cauliflower + Potato - Classic",
    "MI Cauliflower + Potato - Party Pack": "Cauliflower + Potato - Party Pack",
    "MI Chicken Tikka - Catering - Additional Item": "Catering - Additional Item - Chicken Tikka",
    "MI Chicken Tikka - Classic": "Chicken Tikka - Classic",
    "MI Chicken Tikka - Party Pack": "Chicken Tikka - Party Pack",
    "MI Coconut Ginger - Catering - Additional Item": "Catering - Additional Item - Coconut Ginger",
    "MI Coconut Ginger - Classic": "Coconut Ginger - Classic",
    "MI Coconut Ginger - Party Pack": "Coconut Ginger - Party Pack",
    "MI Combo Naan Basket Large": "Combo Naan - Large (serves 20)",
    "MI Combo Naan Basket Small": "Combo Naan - Small (serves 10)",
    "MI Cucumber Cubes - Classic": "Cucumber Cubes - Classic",
    "MI Cucumber Cubes - Party Pack": "Cucumber Cubes - Party Pack",
    "MI Garlic Naan Basket Large": "Garlic Naan - Large (serves 20)",
    "MI HUNGRY Harvest Veggies": "HUNGRY Harvest Vegetables",
    "MI House Vinaigrette Dressing - Classic": "Chili Lime Vinaigrette - Classic",
    "MI House Vinaigrette Dressing - Party Pack": "Chili Lime Vinaigrette - Party Pack",
    "MI Indian Street Corn - Classic": "Indian Street Corn - Classic",
    "MI Indian Street Corn - Party Pack": "Indian Street Corn - Party Pack",
    "MI Kachumber Salad - Classic": "Kachumber Salad - Classic",
    "MI Kachumber Salad - Party Pack": "Kachumber Salad - Party Pack",
    "MI Kokum Punch - 1 Gallon": "Kokum Punch - 1 Gallon",
    "MI Kokum Vinaigrette Dressing - Classic": "Kokum Vinaigrette Dressing - Classic",
    "MI Kokum Vinaigrette Dressing - Party Pack": "Kokum Vinaigrette Dressing - Party Pack",
    "MI Lamb Kebab - Catering - Additional Item": "Catering - Additional Item - Lamb Kebab",
    "MI Lamb Kebab - Classic": "Lamb Kebab - Classic",
    "MI Lamb Kebab - Party Pack": "Lamb Kebab - Party Pack",
    "MI Lemon Turmeric Rice - Catering - Additional Item": "Catering - Additional Item - Lemon Turmeric Rice",
    "MI Lemon Turmeric Rice - Classic": "Lemon Turmeric Rice - Classic",
    "MI Lemon Turmeric Rice - Party Pack": "Lemon Turmeric Rice - Party Pack",
    "MI Mango Salsa - Classic": "Mango Salsa - Classic",
    "MI Mango Salsa - Party Pack": "Mango Salsa - Party Pack",
    "MI Masala Chai Cookies - 10": "10 Chai Cookie Basket",
    "MI Masala Chai Cookies - 100": "100 Chai Cookie Basket",
    "MI Masala Chai Cookies - 20": "20 Chai Cookie Basket",
    "MI Masala Chai Cookies - 30": "30 Chai Cookie Basket",
    "MI Masala Chai Cookies - 50": "50 Chai Cookie Basket",
    "MI Masala Quinoa - Catering - Additional Item": "Catering - Additional Item - Masala Quinoa",
    "MI Masala Quinoa - Classic": "Masala Quinoa - Classic",
    "MI Masala Quinoa - Party Pack": "Masala Quinoa - Party Pack",
    "MI Mint Cardamom Limeade - 1 Gallon": "Mint Cardamom Limeade - 1 Gallon",
    "MI Mint Cilantro Chutney - Classic": "Mint Cilantro Chutney - Classic",
    "MI Mint Cilantro Chutney - Party Pack": "Mint Cilantro Chutney - Party Pack",
    "MI Peanut Sesame - Catering - Additional Item": "Catering - Additional Item - Peanut Sesame",
    "MI Peanut Sesame - Classic": "Peanut Sesame - Classic",
    "MI Peanut Sesame - Party Pack": "Peanut Sesame - Party Pack",
    "MI Pickled Onions - Classic": "Pickled Onions - Classic",
    "MI Pickled Onions - Party Pack": "Pickled Onions - Party Pack",
    "MI Plain Naan Basket Large": "Plain Naan - Large (serves 20)",
    "MI Roasted Lentils - Classic": "Roasted Lentils - Classic",
    "MI Roasted Lentils - Party Pack": "Roasted Lentils - Party Pack",
    "MI Samosa Tray - 10": "Samosa Tray - 10 Pieces",
    "MI Samosa Tray - 100": "Samosa Tray - 100 Pieces",
    "MI Samosa Tray - 20": "Samosa Tray - 20 Pieces",
    "MI Samosa Tray - 50": "Samosa Tray - 50 Pieces",
    "MI Sauteed Spinach - Catering - Additional Item": "Catering - Additional Item - Sauteed Spinach",
    "MI Sexygreens - Classic": "Sexygreens - Classic",
    "MI Sexygreens - Party Pack": "Sexygreens - Party Pack",
    "MI Shredded Paneer - Classic": "Shredded Paneer Cheese - Classic",
    "MI Shredded Paneer - Party Pack": "Shredded Paneer Cheese - Party Pack",
    "MI South Indian Rice Noodles - Catering - Additional Item": "Catering - Additional Item - South Indian Rice Noodles",
    "MI South Indian Rice Noodles - Classic": "South Indian Rice Noodles - Classic",
    "MI South Indian Rice Noodles - Party Pack": "South Indian Rice Noodles - Party Pack",
    "MI Spiced Chickpeas - Catering - Additional Item": "Catering - Additional Item - Spiced Chickpeas",
    "MI Spiced Chickpeas - Classic": "Spiced Chickpeas - Classic",
    "MI Spiced Chickpeas - Party Pack": "Spiced Chickpeas - Party Pack",
    "MI Spicy Chicken Burrito - In House": "Spicy Chicken Burrito",
    "MI Spicy Chili Chicken - Classic": "Spicy Chili Chicken - Classic",
    "MI Spicy Chili Chicken - Party Pack": "Spicy Chili Chicken - Party Pack",
    # NOTE: "MI Sweet Tamarind Chutney" is NOT force-mapped to "Sweet Tamarind -
    # Classic" here (was, until 2026-08-05) -- that recipe's real/natural
    # clean_name is bare "Sweet Tamarind Chutney", a real, heavily-used
    # APP/TPD modifier (see fact_modifiers). Forcing it globally broke that
    # channel's costing to get catering's "Sweet Tamarind - Classic" a cost.
    # Catering's redirect to this same shared recipe lives in the dashboard's
    # CATERING_MOD_ALIAS_CTE instead (mod_alias, lib/queries.ts), which is
    # catering-query-scoped only and can't touch IH/Online.
    "MI Tamarind Chili - Catering - Additional Item": "Catering - Additional Item - Tamarind Chili (Spicy)",
    "MI Tamarind Chili - Classic": "Tamarind Chili (Spicy) - Classic",
    "MI Tamarind Chili - Party Pack": "Tamarind Chili (Spicy) - Party Pack",
    "MI Tamarind Ginger Chutney - Classic": "Ginger Tamarind Chutney - Classic",
    "MI Tamarind Ginger Chutney - Party Pack": "Ginger Tamarind Chutney - Party Pack",
    "MI Tandoori Paneer - Catering - Additional Item": "Catering - Additional Item - Tandoori Paneer",
    "MI Tandoori Paneer - Classic": "Tandoori Paneer - Classic",
    "MI Tandoori Paneer - Party Pack": "Tandoori Paneer - Party Pack",
    "MI Tandoori Paneer Burrito - In House": "Tandoori Paneer Burrito",
    "MI Tikka Masala - Catering - Additional Item": "Catering - Additional Item - Tikka Masala",
    "MI Tikka Masala - Classic": "Tikka Masala - Classic",
    "MI Tikka Masala - Party Pack": "Tikka Masala - Party Pack",
    "MI Toasted Cumin Yogurt - Classic": "Toasted Cumin Yogurt - Classic",
    "MI Toasted Cumin Yogurt - Party Pack": "Toasted Cumin Yogurt - Party Pack",
    "MI Tomato Garlic - Catering - Additional Item": "Catering - Additional Item - Tomato Garlic (Butter Masala)",
    "MI Tomato Garlic - Classic": "Tomato Garlic (Mild) - Classic",
    "MI Tomato Garlic - Party Pack": "Tomato Garlic (Mild) - Party Pack",
    "MI Turmeric Ginger Lemonade - 1 Gallon": "Turmeric Ginger Lemonade - 1 Gallon",
    "MI Vegan Veggie Burrito - In House": "Vegan Veggie Burrito",
    "MI Zerocater Chicken Tikka": "ZeroCater - Chicken Tikka",
    "MI Zerocater Chili Lime Vinaigrette": "ZeroCater - Chili Lime Vinaigrette",
    "MI Zerocater Kachumber Salad": "ZeroCater - Kachumber Salad",
    "MI Zerocater Lamb Kebab Meatballs": "ZeroCater - Lamb Kebab Meatballs",
    "MI Zerocater Lemon Turmeric Rice": "ZeroCater - Lemon Turmeric Rice",
    "MI Zerocater Sautéed Spinach": "ZeroCater - Sautéed Spinach",
    # Resolved conflicts (owner confirmed 2026-08-01): recipe used two different
    # names depending on context — owner picked which one is the recipe's real
    # clean_name; the other name needs the separate modifier-name->recipe lookup
    # (MODIFIER_COST_KNOWN_FIXES.md), not a clean_name override, since it's not
    # this recipe's own identity. Verified zero current IH/Online usage either
    # way — catering-only, safe under the P6 scope restriction.
    "MI Garlic Naan Basket Small": "Garlic Naan - Small (serves 10)",
    # Correction (2026-08-03): "Plain Naan - Small (serves 10)" was originally
    # thought to share MI Garlic Naan Basket Small's recipe — wrong. It's a
    # genuinely separate recipe with its own real cost data every period.
    # Same pattern as the already-correct "MI Plain Naan Basket Large" ->
    # "Plain Naan - Large (serves 20)". Verified zero IH/Online usage.
    "MI Plain Naan Basket Small": "Plain Naan - Small (serves 10)",
    # "Roasted Vegetables" confirmed (owner, 2026-08-03) as the correct clean_name
    # for the whole MI Harvest Veggies family — "Harvest Vegetables ..." was noise/
    # a duplicate in the source sheet. NOTE: "MI Harvest Veggies - Kids" is
    # deliberately NOT included here — "Roasted Vegetables - Kids" is used by the
    # current live Kids Meal item, out of scope under the catering-only restriction.
    "MI Harvest Veggies - Classic": "Roasted Vegetables - Classic",
    "MI Harvest Veggies - Party Pack": "Roasted Vegetables - Party Pack",
    "MI Harvest Vegetables - Catering - Additional Item": "Catering - Additional Item - Roasted Vegetables",
}


def cmd_load_r365_modifier_cost(args: argparse.Namespace) -> None:
    """Load all P*ModifierCost.xlsx from Data/R365Data/ModifierCost/ into analytics.r365_modifier_cost."""
    import openpyxl

    data_dir = _R365_DATA_ROOT / "ModifierCost"
    files = sorted(data_dir.glob("P*ModifierCost.xlsx"))
    if not files:
        raise SystemExit(f"No P*ModifierCost.xlsx files found in {data_dir}")

    conn = db.connect()
    sql_path = Path(__file__).resolve().parents[1] / "sql" / "011_analytics_modifier_cost.sql"
    conn.execute(sql_path.read_text())
    conn.commit()

    # Ground truth for the review-flag check below: every clean_name that has
    # actually been ordered by a real guest, ever. Built once — doesn't change
    # per period file.
    with conn.cursor() as cur:
        cur.execute("SELECT DISTINCT LOWER(canonical_name) FROM public.fact_modifiers")
        real_modifier_names = {r[0] for r in cur.fetchall()}

    for path in files:
        m = re.match(r"P(\d{2})(\d{4})ModifierCost", path.stem, re.IGNORECASE)
        if not m:
            log.warning("skipping %s — can't parse period/year from filename", path.name)
            continue
        period = f"P{m.group(1)}-{m.group(2)}"

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        seen: set[str] = set()
        rows = []
        missing_clean_name = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            recipe_name = str(row[0]).strip() if row[0] else None
            if not recipe_name or recipe_name in seen:
                continue
            seen.add(recipe_name)

            def _cost(val):
                if val is None or str(val) == "#ERROR!":
                    return None
                try:
                    return Decimal(str(val))
                except InvalidOperation:
                    return None

            clean_name = str(row[1]).strip() if len(row) > 1 and row[1] else None
            if not clean_name:
                missing_clean_name += 1
                clean_name = _derive_modifier_clean_name(recipe_name)
            if recipe_name in _KNOWN_CLEAN_NAME_FIXES and clean_name != _KNOWN_CLEAN_NAME_FIXES[recipe_name]:
                log.warning(
                    "r365-modifier-cost: %s — %s: sheet says clean_name=%r, overriding to known-correct %r",
                    path.name, recipe_name, clean_name, _KNOWN_CLEAN_NAME_FIXES[recipe_name],
                )
                clean_name = _KNOWN_CLEAN_NAME_FIXES[recipe_name]

            rows.append((
                period,
                recipe_name,
                clean_name,
                str(row[14]).strip() if row[14] else None,  # portion_unit
                str(row[16]).strip() if row[16] else None,  # cogs_account
                _cost(row[25]),                             # total_cost
                _cost(row[26]),                             # cost_per_portion
            ))

        if missing_clean_name:
            log.warning(
                "r365-modifier-cost: %s — clean_name (column B) missing for %d/%d rows; "
                "derived via TRIM(SUBSTITUTE(A,\"MI \",\"\")) instead",
                path.name, missing_clean_name, len(rows),
            )

        # Flag clean_names that existed in the prior period but vanished here —
        # usually either a naming drift in the new sheet (like "Romaine" quietly
        # becoming "Romaine Lettuce") or a genuine gap in this period's R365
        # export. Either way it's worth a human's eyes before trusting the load.
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT period FROM analytics.r365_modifier_cost")
            existing_periods = [r[0] for r in cur.fetchall() if r[0] != period]

        def _period_key(p: str) -> tuple[int, int]:
            pm = re.match(r"P(\d{2})-(\d{4})", p)
            return (int(pm.group(2)), int(pm.group(1))) if pm else (0, 0)

        prior_recipe_to_clean: dict[str, str] = {}
        prior_periods = [p for p in existing_periods if _period_key(p) < _period_key(period)]
        if prior_periods:
            prior_period = max(prior_periods, key=_period_key)
            current_clean_names = {r[2] for r in rows}  # in-memory parse, not the DB's pre-upsert state
            with conn.cursor() as cur:
                cur.execute("SELECT recipe_name, clean_name FROM analytics.r365_modifier_cost WHERE period = %s", (prior_period,))
                prior_recipe_to_clean = dict(cur.fetchall())
            prior_clean_names = set(prior_recipe_to_clean.values())
            dropped = sorted(prior_clean_names - current_clean_names)
            if dropped:
                preview = ", ".join(dropped[:15]) + (f" ... (+{len(dropped) - 15} more)" if len(dropped) > 15 else "")
                log.warning(
                    "r365-modifier-cost: %s — %d clean_name(s) present in %s but missing here: %s",
                    path.name, len(dropped), prior_period, preview,
                )

        # Flag genuine clean_name drift before this gets trusted — R365's clean_name
        # can be wrong in ways _KNOWN_CLEAN_NAME_FIXES doesn't cover yet (that dict
        # only has cases already found by hand; see MODIFIER_COST_KNOWN_FIXES.md).
        # Only flag a recipe if its clean_name is NEW or CHANGED since the prior
        # period AND the new value has zero real orders — a stable clean_name that's
        # always been low-volume/rare (there are hundreds of those in a menu this
        # size) is not a bug and would just be noise here.
        flagged = []
        for _, recipe_name, clean_name, _, _, _, cost_per_portion in rows:
            rn_upper = recipe_name.upper()
            if rn_upper.startswith("TEST") or rn_upper.startswith("BATCH") or rn_upper.startswith("NEST"):
                continue  # internal prep/test recipes, never customer-facing
            cn_lower = clean_name.lower()
            if cn_lower.startswith("skip ") or cn_lower.startswith("no "):
                continue  # deliberate zero-cost markers (see pc_refresh.sql)
            if prior_recipe_to_clean.get(recipe_name) == clean_name:
                continue  # unchanged since last period — not new drift
            if cn_lower not in real_modifier_names:
                flagged.append((recipe_name, clean_name, cost_per_portion))

        if flagged:
            error_dir = Path(__file__).resolve().parents[1] / "error"
            error_dir.mkdir(exist_ok=True)
            out_path = error_dir / f"{period}_modifier_cost_review.xlsx"
            review_wb = openpyxl.Workbook()
            review_ws = review_wb.active
            review_ws.title = "review"
            review_ws.append(["period", "recipe_name", "clean_name (current)", "cost_per_portion", "correct_clean_name"])
            for recipe_name, clean_name, cost_per_portion in flagged:
                review_ws.append([
                    period, recipe_name, clean_name,
                    float(cost_per_portion) if cost_per_portion is not None else None,
                    None,
                ])
            review_wb.save(out_path)
            log.warning(
                "r365-modifier-cost: %s — %d clean_name(s) have no matching real order; written to %s for review",
                path.name, len(flagged), out_path,
            )

        with conn.cursor() as cur:
            cur.executemany(
                """
                INSERT INTO analytics.r365_modifier_cost
                    (period, recipe_name, clean_name, portion_unit, cogs_account,
                     total_cost, cost_per_portion)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (period, recipe_name) DO UPDATE SET
                    clean_name       = EXCLUDED.clean_name,
                    portion_unit     = EXCLUDED.portion_unit,
                    cogs_account     = EXCLUDED.cogs_account,
                    total_cost       = EXCLUDED.total_cost,
                    cost_per_portion = EXCLUDED.cost_per_portion,
                    loaded_at        = now()
                """,
                rows,
            )
        conn.commit()
        log.info("r365-modifier-cost: %s → %d rows upserted", path.name, len(rows))

    conn.close()


def cmd_load_r365_item_cost(args: argparse.Namespace) -> None:
    """Load all P*ItemCost.xlsx from Data/R365Data/ItemCost/ into analytics.r365_item_cost."""
    import openpyxl

    data_dir = _R365_DATA_ROOT / "ItemCost"
    files = sorted(data_dir.glob("P*ItemCost.xlsx"))
    if not files:
        raise SystemExit(f"No P*ItemCost.xlsx files found in {data_dir}")

    conn = db.connect()
    sql_path = Path(__file__).resolve().parents[1] / "sql" / "012_analytics_item_cost.sql"
    conn.execute(sql_path.read_text())
    conn.commit()

    for path in files:
        m = re.match(r"P(\d{2})(\d{4})ItemCost", path.stem, re.IGNORECASE)
        if not m:
            log.warning("skipping %s — can't parse period/year from filename", path.name)
            continue
        period = f"P{m.group(1)}-{m.group(2)}"

        wb = openpyxl.load_workbook(path)
        ws = wb["Sheet1"]

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            menu      = str(row[0]).strip() if row[0] else None
            item_name = str(row[2]).strip() if row[2] else None
            if not menu or not item_name:
                continue

            raw_cost = str(row[6]).strip().lstrip("$") if row[6] else None
            try:
                avg_cost = Decimal(raw_cost) if raw_cost else None
            except InvalidOperation:
                avg_cost = None

            rows.append((
                period,
                menu,
                item_name,
                _normalize_r365_item_name(str(row[3]).strip()) if row[3] else (
                    _normalize_r365_item_name(item_name)
                ),  # item_name_updated — normalized; falls back to item_name
                str(row[1]).strip() if row[1] else None,  # menu_group
                str(row[4]).strip() if row[4] else None,  # category_1
                str(row[5]).strip() if row[5] else None,  # category_2
                avg_cost,
            ))

        with conn.cursor() as cur:
            cur.executemany(
                """
                INSERT INTO analytics.r365_item_cost
                    (period, menu, item_name, item_name_updated, menu_group,
                     category_1, category_2, avg_cost)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (period, menu, item_name) DO UPDATE SET
                    item_name_updated = EXCLUDED.item_name_updated,
                    menu_group        = EXCLUDED.menu_group,
                    category_1        = EXCLUDED.category_1,
                    category_2        = EXCLUDED.category_2,
                    avg_cost          = EXCLUDED.avg_cost,
                    loaded_at         = now()
                """,
                rows,
            )
        conn.commit()
        log.info("r365-item-cost: %s → %d rows upserted", path.name, len(rows))

    conn.close()


# ---------------------------------------------------------------------------
# R365 item_name_updated normalisation
# ---------------------------------------------------------------------------
_R365_SUFFIX_RE = re.compile(
    r"\s*-\s*(in house|catering|3pd|club feast|ezcater|side|gameday|in-house)\s*$",
    re.IGNORECASE,
)

_VENDOR_PREFIX_RE = re.compile(
    r"^(fooda|hungry|sharebite|aramark|eurest|cureate|zerocater|foodworks|"
    r"metz|territory|wck|catercow|guest\s+services|offsite)\s+",
    re.IGNORECASE,
)

# Approved canonical mappings (raw/variant lowercase → target)
_R365_ITEM_CANONICAL: dict[str, str] = {
    # Cauli variants
    "cauliflower + quinoa":                "Spiced Cauli + Quinoa Bowl",
    "cauliflower + quinoa bowl":           "Spiced Cauli + Quinoa Bowl",
    "sharebite cauliflower + quinoa bowl": "Spiced Cauli + Quinoa Bowl",
    # BYO renames
    "grain bowl":           "BYO Grain Bowl",
    "salad bowl":           "BYO Salad Bowl",
    "greens + grains bowl": "BYO Greens + Grains Bowl",
    "harvest chicken bowl": "BYO Greens + Grains Bowl",
    "kids byo":             "Kids Meal",
    "burrito":              "BYO Indian Burrito",
    # Retail merges (from consolidation review)
    # NOTE: Fooda Chicken Tikka Masala is deliberately NOT merged here — Rahul
    # confirmed it's a distinct item from retail "Chicken Tikka Masala" (see
    # _NAME_CONSOLIDATIONS below). It was wrongly merged here until 2026-08-05,
    # silently colliding item-cost matching with the real Chicken Tikka Masala
    # catering row (both resolved to the same item_name_updated, so the
    # DISTINCT ON "freshest row" pick in the dashboard's base-cost CTE could
    # give either item the other's cost).
    "cureate spicy chili chicken bowl":   "Spicy Chili Chicken Bowl",
    "fooda mango lassi":                  "Mango Lassi",
    "eurest mango lassi":                 "Mango Lassi",
    "hungry byo chicken tikka bowl":      "Chicken Tikka Bowl",
    "sharebite chicken tikka bowl":       "Chicken Tikka Bowl",
    "hungry masala chai cookies":         "Masala Chai Cookies",
    "sharebite spicy chili chicken bowl": "Spicy Chili Chicken Bowl",
}


def _normalize_r365_item_name(raw: str) -> str:
    """Strip known suffixes then apply canonical mapping to item_name_updated."""
    s = raw.strip()
    while True:
        stripped = _R365_SUFFIX_RE.sub("", s).strip()
        if stripped == s:
            break
        s = stripped
    canonical = _R365_ITEM_CANONICAL.get(s.lower())
    if canonical:
        return canonical
    if _VENDOR_PREFIX_RE.match(s):
        log.warning(
            "r365-item-cost: unmapped vendor-prefixed name '%s' — add to "
            "_R365_ITEM_CANONICAL if it should be merged with a retail item",
            s,
        )
    return s


_NAME_CONSOLIDATIONS = [
    # Retail match merges (vendor-prefix → canonical retail name)
    # NOTE: Fooda Chicken Tikka Masala is deliberately NOT merged here — Rahul
    # confirmed it's a distinct item from retail "Chicken Tikka Masala".
    ("Cureate Spicy Chili Chicken Bowl",  "Spicy Chili Chicken Bowl"),
    ("Fooda Mango Lassi",                 "Mango Lassi"),
    ("Eurest Mango Lassi",                "Mango Lassi"),
    ("HUNGRY BYO Chicken Tikka Bowl",     "Chicken Tikka Bowl"),
    ("Sharebite Chicken Tikka Bowl",      "Chicken Tikka Bowl"),
    ("HUNGRY Masala Chai Cookies",        "Masala Chai Cookies"),
    ("Sharebite Spicy Chili Chicken Bowl","Spicy Chili Chicken Bowl"),
    # Sharebite Cauli must go directly to the final renamed value
    ("Sharebite Cauliflower + Quinoa Bowl","Spiced Cauli + Quinoa Bowl"),
    # Short-name renames (run after retail merges so Cauli chain resolves correctly)
    ("Grain Bowl",              "BYO Grain Bowl"),
    ("Greens + Grains Bowl",    "BYO Greens + Grains Bowl"),
    ("Salad Bowl",              "BYO Salad Bowl"),
    ("Cauliflower + Quinoa",    "Spiced Cauli + Quinoa Bowl"),
    ("Cauliflower + Quinoa Bowl","Spiced Cauli + Quinoa Bowl"),
    ("Kids BYO",                "Kids Meal"),
]


def cmd_consolidate_names(args: argparse.Namespace) -> None:
    conn = db.connect()
    total = 0
    with conn.cursor() as cur:
        for raw_name, canonical in _NAME_CONSOLIDATIONS:
            cur.execute(
                "UPDATE public.fact_order_lines SET canonical_name = %s WHERE canonical_name = %s",
                (canonical, raw_name),
            )
            n = cur.rowcount
            if n:
                log.info("consolidate-names: '%s' → '%s' (%d rows)", raw_name, canonical, n)
            total += n
    conn.commit()
    conn.close()
    log.info("consolidate-names: done — %d rows updated", total)


_BACKFILL_OG_SQL = """
WITH raw_mod_og AS (
    -- Extract modifier_guid -> og_guid per location from raw orders (all nesting levels)
    SELECT DISTINCT
        o.location_code,
        mod_elem->>'guid'                       AS modifier_guid,
        (mod_elem->'optionGroup')->>'guid'       AS og_guid
    FROM raw.toast_orders o,
         jsonb_array_elements(o.payload->'checks')      AS check_elem,
         jsonb_array_elements(check_elem->'selections') AS sel_elem,
         jsonb_array_elements(sel_elem->'modifiers')    AS mod_elem
    WHERE (mod_elem->'optionGroup')->>'guid' IS NOT NULL
      AND mod_elem->>'guid' IS NOT NULL
),
raw_og_names AS (
    -- Resolve og_guid -> og_name per location from stored menus config
    SELECT
        c.location_code,
        og_ref.value->>'guid' AS og_guid,
        og_ref.value->>'name' AS og_name
    FROM raw.toast_config c,
         jsonb_each(c.payload->'modifierGroupReferences') AS og_ref
    WHERE c.config_type = 'menus'
      AND og_ref.value->>'guid' IS NOT NULL
      AND og_ref.value->>'name' IS NOT NULL
)
UPDATE public.fact_modifiers fm
SET
    option_group_guid = rmo.og_guid,
    option_group_name = ron.og_name
FROM raw_mod_og rmo
JOIN raw_og_names ron
    ON ron.location_code = rmo.location_code
   AND ron.og_guid = rmo.og_guid
WHERE fm.modifier_guid = rmo.modifier_guid
  AND fm.option_group_guid IS DISTINCT FROM rmo.og_guid
"""


def cmd_backfill_option_groups(args: argparse.Namespace) -> None:
    """Backfill option_group_guid/name on fact_modifiers using raw order history."""
    conn = db.connect()
    with conn.cursor() as cur:
        cur.execute(_BACKFILL_OG_SQL)
        updated = cur.rowcount
    conn.commit()
    conn.close()
    log.info("backfill-option-groups: done — %d fact_modifiers rows updated", updated)


def main() -> None:
    p = argparse.ArgumentParser(prog="toast_pipeline")
    sub = p.add_subparsers(dest="cmd", required=True)

    sub.add_parser("init-db").set_defaults(func=cmd_init_db)
    sub.add_parser("precompute").set_defaults(func=cmd_precompute)

    runp = sub.add_parser("run")
    runp.add_argument("--start")
    runp.add_argument("--end")
    runp.add_argument("--locations", help="comma-separated location codes; default all")
    runp.set_defaults(func=cmd_run)

    rp = sub.add_parser("reparse")
    rp.add_argument("--start", required=True)
    rp.add_argument("--end", required=True)
    rp.add_argument("--locations", help="comma-separated location codes; default all")
    rp.add_argument("--merge", action="store_true", help="merge to public after reparsing")
    rp.set_defaults(func=cmd_reparse)

    sub.add_parser("merge").set_defaults(func=cmd_merge)
    sub.add_parser("validate").set_defaults(func=cmd_validate)
    sub.add_parser("bikky-instore",
                   help="load all P*IS.csv from Data/Bikkydata/InStore/ into public.fact_bikky_instore"
                   ).set_defaults(func=cmd_bikky_instore)
    sub.add_parser("bikky-3pd",
                   help="load all P*Del.csv from Data/Bikkydata/3PD+Loyalty/ into public.fact_bikky_3pd_loyalty"
                   ).set_defaults(func=cmd_bikky_3pd)
    sub.add_parser("load-lookups",
                   help="load LookupItemAndModifierType.xlsx and LookupMenuBreakdown.xlsx into analytics"
                   ).set_defaults(func=cmd_load_lookups)
    sub.add_parser("r365-modifier-cost",
                   help="load all P*ModifierCost.xlsx from Data/R365Data/ModifierCost/ into analytics.r365_modifier_cost"
                   ).set_defaults(func=cmd_load_r365_modifier_cost)
    sub.add_parser("r365-item-cost",
                   help="load all P*ItemCost.xlsx from Data/R365Data/ItemCost/ into analytics.r365_item_cost"
                   ).set_defaults(func=cmd_load_r365_item_cost)
    sub.add_parser("consolidate-names",
                   help="apply approved name consolidations directly to public.fact_order_lines.canonical_name"
                   ).set_defaults(func=cmd_consolidate_names)
    sub.add_parser("backfill-option-groups",
                   help="backfill option_group_guid/name on fact_modifiers from raw order history"
                   ).set_defaults(func=cmd_backfill_option_groups)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
